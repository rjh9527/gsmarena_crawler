import json
import os
import re

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment

from config import *


def store_in_json(data, brand):
    file_path = os.path.join(os.getcwd(), "device_url_data", brand + ".json")
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(json.dumps(data))


def store_brand_page_url(data):
    with open("brand_page_url.json", "w", encoding="utf-8") as f:
        f.write(json.dumps(data))


def read_brand_page_url():
    with open(BRAND_PAGE_URL_FILENAME, "r", encoding="utf-8") as f:
        brand_page_url_map = json.loads(f.read())
    return brand_page_url_map


def read_device_urls(brand):
    file_path = os.path.join(os.getcwd(), "device_url_data", brand + ".json")
    if not os.path.exists(file_path):
        return None

    with open(file_path, "r", encoding="utf-8") as f:
        device_urls = json.loads(f.read())
    return device_urls


def get_brand_url():
    """获取各个品牌对应的url"""
    resp = requests.get(BASE_URL, headers=HEADERS)
    if not resp.ok:
        raise Exception("获取品牌异常")
    content = resp.text
    soup = BeautifulSoup(content, "html.parser")
    all_li = soup.find("div", {"class": "brandmenu-v2 light l-box clearfix"}).find("ul").findAll("li")
    # 存储每个手机品牌的url地址
    # brand_url_map = [os.path.join(TARGET_URL, li.a.get("href")) for li in all_li]
    brand_url_map = {
        li.get_text().lower(): os.path.join(BASE_URL, li.a.get("href"))
        for li in all_li
    }
    return brand_url_map


def get_brand_page_template(brand, url):
    """拿到页面url模板,及最大页数"""
    try:
        resp = requests.get(url, headers=HEADERS)
        if not resp.ok:
            return {
                "brand": brand,
                "page_url": None,
                "max_page_num": 0
            }
        content = resp.text
        soup = BeautifulSoup(content, "html.parser")
        all_a = soup.find("div", {"class": "nav-pages"})

        brand_page_url = url
        max_page = 1

        if all_a and len(all_a) > 0:
            all_a = all_a.findAll("a")
            # 存储每个品牌的每个分页的url地址
            brand_page_urls = [os.path.join(BASE_URL, a.get("href")) for a in all_a if "php" in a.get("href")]
            pattern = re.compile(r'-p(\d+)')
            for page_url in brand_page_urls:
                match = pattern.search(page_url)
                if match:
                    if int(match.group(1)) > max_page:
                        max_page = int(match.group(1))
                    start_index = match.start()
                    end_index = match.end()
                    brand_page_url = page_url[:start_index + 2] + "{}" + page_url[end_index:]
        else:
            brand_page_url = url
        return {
            "brand": brand,
            "page_url": brand_page_url,
            "max_page_num": max_page
        }
    except:
        error_msg = "{}品牌获取模板url异常".format(brand)
        print(error_msg)
        return {
            "brand": brand,
            "page_url": None,
            "max_page_num": 0
        }


def get_page_device_url(page_url):
    """获取每一页的设备信息"""
    resp = requests.get(page_url, headers=HEADERS)
    if not resp.ok:
        error_msg = "{}获取异常".format(page_url)
        print(error_msg)
        return None
    content = resp.content
    soup = BeautifulSoup(content, "html.parser")
    all_li = soup.find("div", {"class": "makers"}).find("ul").findAll("li")
    # 存储每个手机品牌的url地址
    device_url_list = [os.path.join(BASE_URL, li.a.get("href")) for li in all_li]
    return device_url_list


def get_device_info(device_url, head_dict):
    """获取设备信息"""
    try:
        resp = requests.get(device_url, headers=HEADERS)
        if not resp.ok:
            error_msg = "{}获取异常".format(device_url)
            print(error_msg)
            return None, None
        content = resp.content
        soup = BeautifulSoup(content, "html.parser")
        div = soup.find("div", {"class": "main main-review right l-box col"})
        brand_str, product_str = div.find("h1", {"class": "specs-phone-name-title"}).text.split(" ", 1)
        specs_info = {"brand": brand_str, "product": product_str, "url": device_url}
        specs_list = div.find("div", {"id": "specs-list"}).findAll("table")
        head = None
        for specs in specs_list:
            tr_list = specs.findAll("tr")
            num = 1
            for tr in tr_list:
                th = tr.find("th")
                if th is not None:
                    head = th.text
                    if head not in head_dict:
                        head_dict[head] = []
                tds = tr.findAll("td")
                if len(tds) <= 0:
                    continue
                key, val = tds[0].text, tds[1].text
                if key == "\xa0":
                    key = "extra_" + head + str(num)
                    num += 1
                specs_info[key] = val
                if head is not None:
                    if key not in head_dict[head]:
                        head_dict[head].append(key)
        return head_dict, specs_info
    except:
        return None, None


def create_workbook(filename, head_dict):
    filename = filename + ".xlsx"
    file_path = os.path.join(os.getcwd(), "device_data", filename)
    # 为每个列表重新排序，将带有 extra_ 前缀的元素移到列表末尾
    for key, values in head_dict.items():
        # 分离出以 'extra_' 开头的和不以 'extra_' 开头的元素
        extra_items = [item for item in values if item.startswith('extra_')]
        non_extra_items = [item for item in values if not item.startswith('extra_')]

        # 将非 extra_ 元素与 extra_ 元素合并，并更新字典的值
        head_dict[key] = non_extra_items + extra_items

    # 创建workbook和worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    # 创建一个居中对齐的样式
    align_center = Alignment(horizontal='center', vertical='center')

    # 将key和value作为标题写入excel
    column = 1  # 起始列
    for key, value in head_dict.items():
        row = 1  # 起始行
        cell = ws.cell(row=row, column=column, value=key)  # 写入key
        cell.alignment = align_center  # 居中
        ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + len(value) - 1)  # 合并单元格
        row += 1  # 下一行
        for sub_title in value:
            sub_cell = ws.cell(row=row, column=column, value=sub_title)  # 写入value
            column += 1  # 下一列

    ws.cell(row=2, column=column, value="url")  # 写入key

    # 保存文件
    wb.save(file_path)


def input_workbook(filename, datas):
    filename = filename + ".xlsx"
    file_path = os.path.join(os.getcwd(), "device_data", filename)
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 选择要操作的工作表
    worksheet = workbook.active

    # 根据每列第一行的值来写入相应的数据
    start_row = worksheet.max_row + 1
    for data in datas:
        start_col = 1
        for col in worksheet.iter_cols(min_row=2):
            # 获取单元格的值
            col_name = col[0].value
            if col_name in data:
                worksheet.cell(row=start_row, column=start_col, value=data[col_name])  # 写入value
            start_col += 1
        start_row += 1

    # 保存修改后的 Excel 文件
    workbook.save(file_path)


if __name__ == '__main__':
    head_dict, specs_info = get_device_info("https://www.gsmarena.com/samsung_z-6403.php", {"commodity": ["brand", "product"]})
    print(specs_info)
